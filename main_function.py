import argparse
import os, sys
import os.path as osp
import torchvision
import numpy as np
import torch
import torch.nn as nn
import torch.optim as optim
import network2, loss
import random, pdb, math, copy
from loss import CrossEntropyLabelSmooth
from sklearn.metrics import confusion_matrix
from actuator_dataset1 import get_dataset, get_dataset_all
from openpyxl import load_workbook
from vis import draw_tsne, draw_confusion_matrix
import torch.nn.functional as F
import warnings
warnings.filterwarnings("ignore")


def op_copy(optimizer):
    for param_group in optimizer.param_groups:
        param_group['lr0'] = param_group['lr']
    return optimizer


def lr_scheduler(optimizer, iter_num, max_iter, gamma=10, power=0.75):
    decay = (1 + gamma * iter_num / max_iter) ** (-power)
    for param_group in optimizer.param_groups:
        param_group['lr'] = param_group['lr0'] * decay
        param_group['weight_decay'] = 1e-3
        param_group['momentum'] = 0.9
        param_group['nesterov'] = True
    return optimizer

def data_load(args):

    dset_loaders = {}
    if args.dataroot not in ['low','high']:
        src_dataset = os.path.join(args.dataroot, args.source_domain)
        tgt_dataset = os.path.join(args.dataroot, args.target_domain)


        dset_loaders["source_tr"], dset_loaders["source_te"], classes = get_dataset(src_dataset,
                                                                                    batch_size=args.batch_size,
                                                                                    randomstate=args.randomstate)
        dset_loaders["target"], dset_loaders["test"], _ = get_dataset(tgt_dataset,
                                                                            batch_size=args.batch_size,
                                                                            randomstate=args.randomstate)
    else:
        if args.source_domain == 'A1':
            source_domain1 = r'D:\data\NASA_EMA_Low_\0\第一次数据'  # Please change this file
        if args.source_domain == 'A2':
            source_domain1 = r'D:\data\NASA_EMA_Low_\1\第一次数据' # Please change this file
        if args.source_domain == 'B1':
            source_domain1 = r'D:\data\第二次试验用的高采样频率数据\0\第一次' # Please change this file
        if args.source_domain == 'B2':
            source_domain1 = r'D:\data\第二次试验用的高采样频率数据\1\第一次' # Please change this file

        if args.target_domain == 'A1':
            target_domain1 = r'D:\data\NASA_EMA_Low_\0\第一次数据' # Please change this file
        if args.target_domain == 'A2':
            target_domain1 = r'D:\data\NASA_EMA_Low_\1\第一次数据' # Please change this file
        if args.target_domain == 'B1':
            target_domain1 = r'D:\data\第二次试验用的高采样频率数据\0\第一次' # Please change this file
        if args.target_domain == 'B2':
            target_domain1 = r'D:\data\第二次试验用的高采样频率数据\1\第一次' # Please change this file

        src_dataset = source_domain1
        tgt_dataset = target_domain1
        dset_loaders["source_tr"], dset_loaders["source_te"], classes = get_dataset_all(src_dataset,
                                                                                    batch_size=args.batch_size,
                                                                                    randomstate=args.randomstate)
        dset_loaders["target"], dset_loaders["test"], _ = get_dataset_all(tgt_dataset,
                                                                            batch_size=args.batch_size,
                                                                            randomstate=args.randomstate)
    return dset_loaders


def cal_acc(loader, netF, netB, netC, t=0, flag=False):
    start_test = True
    with torch.no_grad():
        iter_test = iter(loader)
        for i in range(len(loader)):
            data = iter_test.next()
            inputs = data[0]
            labels = data[1]
            inputs = inputs.cuda()
            outputs = netC(netB(netF(inputs), t=t)[0])
            if start_test:
                all_output = outputs.float().cpu()
                all_label = labels.float()
                start_test = False
            else:
                all_output = torch.cat((all_output, outputs.float().cpu()), 0)
                all_label = torch.cat((all_label, labels.float()), 0)

    # all_output = nn.Softmax(dim=1)(all_output)
    _, predict = torch.max(all_output, 1)
    accuracy = torch.sum(torch.squeeze(predict).float() == all_label).item() / float(all_label.size()[0])
    mean_ent = torch.mean(loss.Entropy(nn.Softmax(dim=1)(all_output))).cpu().data.item()

    return accuracy * 100, mean_ent


def train_source(args):
    dset_loaders = data_load(args) # ----------------
    print('----------------------Source----------------------------------')
    netF = network2.LeNetBase().cuda() # ------------

    netB = network2.feat_bootleneck(type=args.classifier, feature_dim=netF.in_features,
                                    bottleneck_dim=args.bottleneck).cuda() # --------------
    netC = network2.feat_classifier(type=args.layer, class_num=args.class_num, bottleneck_dim=args.bottleneck).cuda() #----------

    args.out_file.write(print_args(netF) + '\n')
    args.out_file.flush()
    args.out_file.write(print_args(netB) + '\n')
    args.out_file.flush()
    args.out_file.write(print_args(netC) + '\n')
    args.out_file.flush()


    param_group = []
    learning_rate = args.lr
    for k, v in netF.named_parameters():
        param_group += [{'params': v, 'lr': learning_rate * 0.1}]
    for k, v in netB.named_parameters():
        param_group += [{'params': v, 'lr': learning_rate}]
    for k, v in netC.named_parameters():
        param_group += [{'params': v, 'lr': learning_rate}]

    optimizer = optim.SGD(param_group)
    optimizer = op_copy(optimizer)  # lr - lr0

    acc_init = 0
    acc_t_best = 0
    max_iter = args.max_epoch * len(dset_loaders["source_tr"])
    interval_iter = max_iter // 10
    iter_num = 0



    netF.train()
    netB.train()
    netC.train()

    smax = 100
    # while iter_num < max_iter:
    for epoch in range(args.max_epoch):
        iter_source = iter(dset_loaders["source_tr"])
        for batch_idx, (inputs_source,
                        labels_source,_) in enumerate(iter_source):

            if inputs_source.size(0) == 1:
                continue
            iter_num += 1

            progress_ratio = batch_idx / (len(dset_loaders) - 1)
            s = (smax - 1 / smax) * progress_ratio + 1 / smax
            lr_scheduler(optimizer, iter_num=iter_num, max_iter=max_iter)

            inputs_source, labels_source = inputs_source.cuda(), labels_source.cuda()
            feature_src, masks = netB(netF(inputs_source), t=0, s=100, all_mask=True) # ------------

            # sparsity regularization for domain attention
            reg = 0
            count = 0
            for m in masks[0]:
                reg += m.sum()  # numerator
                count += np.prod(m.size()).item()  # denominator
            for m in masks[1]:
                reg += m.sum()  # numerator
                count += np.prod(m.size()).item()  # denominator
            reg /= count

            outputs_source1 = netC(feature_src[0])
            outputs_source2 = netC(feature_src[1])
            classifier_loss = CrossEntropyLabelSmooth(
                num_classes=args.class_num, epsilon=args.smooth)(
                outputs_source1, labels_source) + CrossEntropyLabelSmooth(
                num_classes=args.class_num, epsilon=args.smooth)(
                outputs_source2, labels_source) + 0.75 * reg

            optimizer.zero_grad()
            classifier_loss.backward()

            # gradient compensation for embedding layer
            for n, p in netB.em.named_parameters():
                num = torch.cosh(
                    torch.clamp(s * p.data, -10, 10)) + 1
                den = torch.cosh(p.data) + 1
                p.grad.data *= smax / s * num / den
            torch.nn.utils.clip_grad_norm(netF.parameters(), 10000)

            optimizer.step()

        # if iter_num % interval_iter == 0 or iter_num == max_iter:

        if iter_num % interval_iter == 0 or iter_num == max_iter:
            netF.eval()
            netB.eval()
            netC.eval()
            acc_s_tr, _ = cal_acc(dset_loaders['source_tr'], netF, netB, netC, t=0, flag=True)
            acc_s_te, acc_list = cal_acc(dset_loaders['source_te'], netF, netB, netC, t=0, flag=True)
            log_str = 'Task: {}, Iter:{}/{}; Accuracy = {:.2f}%/{:.2f}%'.format(args.name_src, iter_num, max_iter, acc_s_tr, acc_s_te)
            # log_str = 'Task: {}, Iter:{}/{}; Accuracy = {:.2f}%'.format(args.name_src, iter_num, max_iter,
            #                                                             acc_s_te)
            args.out_file.write(log_str + '\n')
            args.out_file.flush()
            print(log_str + '\n')

            acc_t_te, _ = cal_acc(dset_loaders['test'], netF, netB, netC, t=1, flag=True)

            if acc_t_te >= acc_t_best:
                acc_t_best = acc_t_te

            if acc_s_te >= acc_init:
                acc_init = acc_s_te
                best_netF = netF.state_dict()
                best_netB = netB.state_dict()
                best_netC = netC.state_dict()



            netF.train()
            netB.train()
            netC.train()
    print('---------------',acc_t_best)

    torch.save(best_netF, osp.join(args.output_dir_src, "source_F.pt"))
    torch.save(best_netB, osp.join(args.output_dir_src, "source_B.pt"))
    torch.save(best_netC, osp.join(args.output_dir_src, "source_C.pt"))

    return netF, netB, netC, acc_s_tr, acc_s_te, acc_init, acc_t_best


def train_target(args):
    print('----------------------Target----------------------------------')
    dset_loaders = data_load(args)

    netF = network2.LeNetBase().cuda() # ------------

    netB = network2.feat_bootleneck(type=args.classifier, feature_dim=netF.in_features,
                                    bottleneck_dim=args.bottleneck).cuda() # --------------
    netC = network2.feat_classifier(type=args.layer, class_num=args.class_num, bottleneck_dim=args.bottleneck).cuda() #----------

    args.out_file.write(print_args(netF) + '\n')
    args.out_file.flush()
    args.out_file.write(print_args(netB) + '\n')
    args.out_file.flush()
    args.out_file.write(print_args(netC) + '\n')
    args.out_file.flush()


    modelpath = args.output_dir_src + '/source_F.pt'
    netF.load_state_dict(torch.load(modelpath))
    modelpath = args.output_dir_src + '/source_B.pt'
    netB.load_state_dict(torch.load(modelpath))
    modelpath = args.output_dir_src + '/source_C.pt'
    netC.load_state_dict(torch.load(modelpath))
    netC.eval()
    for k, v in netC.named_parameters():
        v.requires_grad = False

    param_group = []
    for k, v in netF.named_parameters():
        # if k.find('bn')!=-1: #
        param_group += [{'params': v, 'lr': args.lr * 0.1}]

    for k, v in netB.named_parameters():
        param_group += [{'params': v, 'lr': args.lr * 0.1}]


    optimizer = optim.SGD(param_group)
    optimizer = op_copy(optimizer)

    #building feature bank and score bank
    loader = dset_loaders["target"]
    num_sample = len(loader.dataset)
    fea_bank = torch.randn(num_sample,256)
    score_bank = torch.randn(num_sample, 4).cuda()

    netF.eval()
    netB.eval()

    with torch.no_grad():
        iter_test = iter(loader)
        for i in range(len(loader)):
            data = iter_test.next()
            inputs = data[0]
            indx = data[-1]
            # print(indx)
            inputs = inputs.cuda()
            output, _ = netB(netF(inputs), t=1)  # a^t
            output_norm = F.normalize(output)
            outputs = netC(output)
            outputs = nn.Softmax(-1)(outputs)
            fea_bank[indx] = output_norm.detach().clone().cpu()
            score_bank[indx] = outputs.detach().clone()  #.cpu()


    max_iter = args.max_epoch * len(dset_loaders["target"])
    interval_iter = max_iter // 10
    iter_num = 0

    acc_t_te_be = 0
    acc_s_te_be = 0

    netF.train()
    netB.train()
    acc_log=0

    while iter_num < max_iter:
        try:
            inputs_test, _, tar_idx = iter_test.next()
        except:
            iter_test = iter(dset_loaders["target"])
            inputs_test, _, tar_idx = iter_test.next()


        if inputs_test.size(0) == 1:
            continue

        inputs_test = inputs_test.cuda()

        iter_num += 1
        lr_scheduler(optimizer, iter_num=iter_num, max_iter=max_iter)
        features_test, masks = netB(netF(inputs_test), t=1)
        masks_old = masks
        # print(len(masks[0]))

        outputs_test = netC(features_test)
        softmax_out = nn.Softmax(dim=1)(outputs_test)
        # print('1----------1',softmax_out)
        output_re = softmax_out.unsqueeze(1)
        # print('2----------2', output_re)


        with torch.no_grad():
            output_f_norm = F.normalize(features_test)
            fea_bank[tar_idx].fill_(-0.1)    #do not use the current mini-batch in fea_bank
            output_f_= output_f_norm.cpu().detach().clone()
            distance = output_f_@fea_bank.T
            _, idx_near = torch.topk(distance,
                                    dim=-1,
                                    largest=True,
                                    k = 10)
            score_near = score_bank[idx_near]    #batch x K x num_class
            score_near=score_near.permute(0,2,1)

            # update banks
            fea_bank[tar_idx] = output_f_.detach().clone().cpu()
            score_bank[tar_idx] = softmax_out.detach().clone()  #.cpu()

        const = torch.log(torch.bmm(output_re,score_near)).sum(-1)
        loss = -torch.mean(const)

        msoftmax = softmax_out.mean(dim=0)
        gentropy_loss = torch.sum(msoftmax *
                                    torch.log(msoftmax + args.epsilon))

        loss += gentropy_loss

        optimizer.zero_grad()
        loss.backward()

        for n, p in netB.bottleneck.named_parameters():
            if n.find('bias') == -1:
                # print(((1 - masks_old)).view(-1, 1))
                mask_ = ((1 - masks_old)).view(-1, 1).expand(256, 512).cuda()
                # print(mask_)
                # print(p.grad.data)
                p.grad.data *= mask_
            else:  #no bias here
                mask_ = ((1 - masks_old)).squeeze().cuda()
                p.grad.data *= mask_

        # for n, p in netC.named_parameters():
        #     if n.find('weight_v') != -1:
        #         masks__=masks_old.view(1,-1).expand(4,256)
        #         mask_ = ((1 - masks__)).cuda()
        #         p.grad.data *= mask_

        for n, p in netB.bn.named_parameters():
            mask_ = ((1 - masks_old)).view(-1).cuda()
            p.grad.data *= mask_

        optimizer.step()

        if iter_num % interval_iter == 0 or iter_num == max_iter:
            netF.eval()
            netB.eval()
            netC.eval()

            acc_t_tr_last, _ = cal_acc(dset_loaders['target'], netF, netB, netC,t=1,flag= True)
            acc_s_tr_last, _ = cal_acc(dset_loaders['source_tr'], netF, netB, netC,t=0,flag= True)

            acc, acc_list = cal_acc(dset_loaders['test'], netF, netB, netC,t=1,flag= True)
            acc11, accS_list = cal_acc(dset_loaders['source_te'], netF, netB, netC, t=0, flag=True)
            if acc > acc_t_te_be:
                acc_t_te_be = acc
                acc_s_te_be = acc11
                best_netF = netF.state_dict()
                best_netB = netB.state_dict()
                best_netC = netC.state_dict()
                draw_netF = netF
                draw_netB = netB
                draw_netC = netC


            log_str = 'Task: {}, Iter:{}/{}; Accuracy on target = {:.2f}%, Accuracy on source = {:.2f}%'.format(
                args.name, iter_num, max_iter, acc, acc11)

            # if acc11 > acc_s_te_be:
            #     acc_s_te_be = acc11

            args.out_file.write(log_str + '\n')
            args.out_file.flush()
            print(log_str + '\n')
            netF.train()
            netB.train()
            # netC.train()

    if args.issave:
        torch.save(best_netF, osp.join(args.output_dir_src, "target_F_" + 'final' + ".pt"))
        torch.save(best_netB, osp.join(args.output_dir_src, "target_B_" + 'final' + ".pt"))
        torch.save(best_netC, osp.join(args.output_dir_src, "target_C_" + 'final' + ".pt"))
        # title = 'A'
        # draw_confusion_matrix(draw_netF, draw_netB, draw_netC, dset_loaders['test'], args.output_dir_src, iter_num, title)
        # draw_tsne(draw_netF, draw_netB, draw_netC, dset_loaders['source_te'], dset_loaders['test'], args.output_dir_src,
        #           iter_num, title, separate=True)
        # draw_tsne(draw_netF, draw_netB, draw_netC, dset_loaders['source_te'], dset_loaders['test'], args.output_dir_src,
        #           iter_num, title, separate=False)
    acc_t_te_last = acc
    acc_t_te_best = acc_t_te_be
    acc_s_te_last = acc11
    acc_s_te_best = acc_s_te_be

    return netF, netB, netC, acc_t_tr_last, acc_t_te_last, acc_t_te_best, acc_s_tr_last, acc_s_te_last, acc_s_te_best

def test_target(args):
    dset_loaders = data_load(args)
    ## set base network

    netF = network2.LeNetBase().cuda() # ------------

    netB = network2.feat_bootleneck(type=args.classifier, feature_dim=netF.in_features,
                                    bottleneck_dim=args.bottleneck).cuda() # --------------
    netC = network2.feat_classifier(type=args.layer, class_num=args.class_num, bottleneck_dim=args.bottleneck).cuda()

    args.modelpath = args.output_dir_src + '/source_F.pt'
    netF.load_state_dict(torch.load(args.modelpath))
    args.modelpath = args.output_dir_src + '/source_B.pt'
    netB.load_state_dict(torch.load(args.modelpath))
    args.modelpath = args.output_dir_src + '/source_C.pt'
    netC.load_state_dict(torch.load(args.modelpath))
    netF.eval()
    netB.eval()
    netC.eval()

    acc, acc_list = cal_acc(dset_loaders['test'], netF, netB, netC, t=1, flag=True)
    log_str = '\nTraining: {}, Task: {}, Accuracy = {:.2f}%'.format(args.trte, args.name, acc)

    args.out_file.write(log_str)
    args.out_file.flush()
    print(log_str)
    return acc


def print_args(args):
    s = "==========================================\n"
    for arg, content in args.__dict__.items():
        s += "{}:{}\n".format(arg, content)
    return s


if __name__ == "__main__":


    # # 选择1
    # dataroot_high_0 = r'H:\第6篇\data\第二次试验用的高采样频率数据\0\第一次'
    # dataroot_high_1 = r'H:\第6篇\data\第二次试验用的高采样频率数据\1\第一次'
    #
    # dataroot_high= 'high'

    # 选择3
    dataroot_low_0 = r'D:\第8篇\Github' # Please change this file
    dataroot_low_1 = r'D:\第8篇\Github' # Please change this file

    # 选择4
    dataroot_low = 'low'

    parser = argparse.ArgumentParser(description='GSFDA')
    parser.add_argument('--dataroot', required=False, default=dataroot_low_0, help='dataroot') #
    parser.add_argument('--gpu_id', type=str, nargs='?', default='0', help="device id to run")
    parser.add_argument('--s', type=int, default=0, help="source")
    parser.add_argument('--t', type=int, default=1, help="target")
    parser.add_argument('--max_epoch', type=int, default=30, help="max iterations")
    parser.add_argument('--batch_size', type=int, default=64, help="batch_size")
    parser.add_argument('--worker', type=int, default=4, help="number of workers")
    parser.add_argument('--dset', type=str, default='low_single', choices=['low_single','high_single','low','high'])
    parser.add_argument('--lr', type=float, default=0.01, help="learning rate")
    parser.add_argument('--seed', type=int, default=2020, help="random seed")
    parser.add_argument('--bottleneck', type=int, default=128) # -------------
    parser.add_argument('--epsilon', type=float, default=1e-5)
    parser.add_argument('--layer', type=str, default="wn", choices=["linear", "wn"])
    parser.add_argument('--classifier', type=str, default="bn", choices=["ori", "bn"])
    parser.add_argument('--smooth', type=float, default=0.1)
    parser.add_argument('--output', type=str, default='.\ckps_digits')
    parser.add_argument('--da', type=str, default='uda')
    parser.add_argument('--issave', type=bool, default=True)
    parser.add_argument('--trte', type=str, default='val', choices=['full', 'val'])

    parser.add_argument('--randomstate', required=False, type=int, default=10, help='')
    parser.add_argument('--source_domain',type=str,default='1',choices=['1', '2', '3', '4','A1','A2','B1','B2']) #
    parser.add_argument('--target_domain', type=str, default='2', choices=['1', '2', '3', '4','A1','A2','B1','B2']) #
    parser.add_argument('--file_xlsx', type=str, default='result.xlsx', help='the index of file_xlsx') #
    parser.add_argument('--Index_number', type=int, default=1, help='the number of Index_number')  #

    args = parser.parse_args()



    names = ['train', 'validation']
    args.name_src = names[args.s][0].upper()

    args.class_num = 4
    os.environ["CUDA_VISIBLE_DEVICES"] = args.gpu_id
    SEED = args.seed
    torch.manual_seed(SEED)
    torch.cuda.manual_seed(SEED)
    np.random.seed(SEED)
    random.seed(SEED)
    # torch.backends.cudnn.deterministic = True

    # folder = './data/'
    # args.s_dset_path = folder + args.dset + '/' + names[args.s] + '_list.txt'
    # args.test_dset_path = folder + args.dset + '/' + names[args.t] + '_list.txt'

    # args.output_dir_src = osp.join(args.output, args.dset)
    args.output_dir_src = osp.join(args.output, 'seed' + str(args.seed), args.dset + 'S_' + args.source_domain + '_'+ 'T_'
                               + args.target_domain + '_' + 'I_' + str(args.Index_number))


    if not osp.exists(args.output_dir_src):
        os.mkdir(args.output_dir_src)

    if not osp.exists(osp.join(args.output_dir_src + '/source_F.pt')):
        args.out_file = open(osp.join(args.output_dir_src, 'log.txt'), 'w')
        args.out_file.write(print_args(args) + '\n')
        args.out_file.flush()
        netF, netB, netC, acc_s_tr, acc_s_te, acc_s_best, acc_t_best = train_source(args)
        args.name = names[args.s][0].upper() + names[args.t][0].upper()
        acc_t_last = test_target(args)

    args.out_file = open(osp.join(args.output_dir_src, 'log_target' + '.txt'), 'w')
    args.out_file.write(print_args(args) + '\n')
    args.out_file.flush()
    _, _, _, acc_t_tr_last, acc_t_te_last, acc_t_te_best, acc_s_tr_last, acc_s_te_last, acc_s_te_best = train_target(args)

    wb = load_workbook(args.file_xlsx)  # 生成一个已存在的wookbook对象
    wb1 = wb.active  # 激活sheet
    wb1.cell(args.Index_number, 1, args.output_dir_src)

    wb1.cell(args.Index_number, 2, args.Index_number)
    wb1.cell(args.Index_number, 3, args.source_domain)
    wb1.cell(args.Index_number, 4, args.target_domain)

    wb1.cell(args.Index_number, 7, acc_s_tr)  #
    wb1.cell(args.Index_number, 8, acc_s_te)
    wb1.cell(args.Index_number, 9, acc_s_best)
    wb1.cell(args.Index_number, 10, acc_t_last)
    wb1.cell(args.Index_number, 11, acc_t_best)
    wb1.cell(args.Index_number, 12, acc_t_tr_last)
    wb1.cell(args.Index_number, 13, acc_t_te_last)
    wb1.cell(args.Index_number, 14, acc_t_te_best)
    wb1.cell(args.Index_number, 15, acc_s_tr_last)
    wb1.cell(args.Index_number, 16, acc_s_te_last)
    wb1.cell(args.Index_number, 17, acc_s_te_best)
    wb.save(args.file_xlsx)

